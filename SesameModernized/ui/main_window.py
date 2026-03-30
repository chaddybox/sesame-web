from __future__ import annotations

import re
import traceback
from dataclasses import asdict, is_dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

from PySide6.QtCore import QSize, Qt
from PySide6.QtGui import QAction, QFont, QIcon, QPixmap
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QVBoxLayout,
    QWidget,
)

from SesameModernized.models.estimator import (
    FitResult,
    ScreeningConfig,
    ScreeningResult,
    SesameEstimator,
)
from SesameModernized.models.nutrient_catalog import (
    NUTRIENT_GROUPS,
    NUTRIENT_OPTIONS_BY_COLUMN,
    PRESETS,
)


APP_VERSION = "1.5"


PROBLEMATIC_SELECTION_RULES = [
    (
        {"dRUP_prot", "dLys_RUP"},
        "Digestible RUP protein and digestible RUP lysine can be strongly collinear because one contains the other.",
    ),
    (
        {"dRUP_prot", "dMet_RUP"},
        "Digestible RUP protein and digestible RUP methionine can be strongly collinear because one contains the other.",
    ),
    (
        {"dRUP_prot", "dHis_RUP"},
        "Digestible RUP protein and digestible RUP histidine can be strongly collinear because one contains the other.",
    ),
    (
        {"dRUP_prot", "dLeu_RUP"},
        "Digestible RUP protein and digestible RUP leucine can be strongly collinear because one contains the other.",
    ),
    (
        {"dRUP_prot", "dIle_RUP"},
        "Digestible RUP protein and digestible RUP isoleucine can be strongly collinear because one contains the other.",
    ),
    (
        {"dRUP_prot", "dVal_RUP"},
        "Digestible RUP protein and digestible RUP valine can be strongly collinear because one contains the other.",
    ),
    (
        {"Total_Fatty_Acids", "C18_1_cis_DM"},
        "TFAs (% DM) and oleic acid (% DM) are biologically related and may be highly correlated in some libraries.",
    ),
    (
        {"NDF", "NDFd"},
        "NDF and Digestible NDF (% DM) often move together and can create unstable fits in smaller datasets.",
    ),
]


def _row_as_dict(r):
    if isinstance(r, dict):
        return r
    if is_dataclass(r):
        return asdict(r)
    return {
        "name": getattr(r, "name", ""),
        "actual_per_t": getattr(r, "actual_per_t", None),
        "predicted_per_t": getattr(r, "predicted_per_t", None),
        "residual": getattr(r, "residual", None),
        "leverage": getattr(r, "leverage", None),
        "student_residual": getattr(r, "student_residual", None),
        "abs_student_residual": getattr(r, "abs_student_residual", None),
        "final_weight": getattr(r, "final_weight", None),
        "ci75_lo": getattr(r, "ci75_lo", None),
        "ci75_hi": getattr(r, "ci75_hi", None),
    }


class MainWindow(QMainWindow):
    def __init__(self, app_icon_path: Optional[str] = None):
        super().__init__()

        self.setWindowTitle(f"Sesame — Modernized v{APP_VERSION}")
        self._assets = Path(__file__).resolve().parent.parent / "assets"
        self._project_root = Path(__file__).resolve().parents[2]
        self._manual_selection_syncing = False
        self._nutrient_checkboxes: Dict[str, QCheckBox] = {}

        if app_icon_path and Path(app_icon_path).exists():
            self.setWindowIcon(QIcon(app_icon_path))
        else:
            icon_path = self._assets / "sesame_icon_multi.ico"
            if icon_path.exists():
                self.setWindowIcon(QIcon(str(icon_path)))

        self._build_fonts()
        self._build_widgets()
        self._build_layout()
        self._build_menus()

        self._estimator = SesameEstimator()
        self._last_dir = str(self._project_root / "data" / "raw")

        self.resize(1180, 780)
        self._apply_global_styles()
        self._apply_preset_selection(0)
        self._refresh_checkbox_enabled_state()
        self._refresh_selection_help()
        self._refresh_selection_summary()
        self._refresh_current_run_panel()

    def _build_fonts(self):
        self.title_font = QFont()
        self.title_font.setPointSize(16)
        self.title_font.setBold(True)

        self.subtitle_font = QFont()
        self.subtitle_font.setPointSize(11)
        self.subtitle_font.setBold(True)

        self.header_font = QFont()
        self.header_font.setPointSize(12)
        self.header_font.setBold(True)

        self.body_font = QFont()
        self.body_font.setPointSize(11)

        self.small_font = QFont()
        self.small_font.setPointSize(10)

    def _build_widgets(self):
        self.preset_box = QComboBox()
        self._presets = [dict(p) for p in PRESETS]
        self._append_user_selected_preset()

        for preset in self._presets:
            self.preset_box.addItem(preset["label"])
        self.preset_box.currentIndexChanged.connect(self._on_preset_changed)
        self.preset_box.setMinimumHeight(36)

        self.preset_label = QLabel("Preset:")
        self.preset_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.preset_label.setFont(self.header_font)

        self.banner = QLabel()
        self.banner.setAlignment(Qt.AlignCenter)
        self._banner_pixmap: QPixmap | None = None
        banner_path = self._assets / "sesame_banner.png"
        if banner_path.exists():
            pm = QPixmap(str(banner_path))
            if not pm.isNull():
                self._banner_pixmap = pm
                self.banner.setPixmap(pm)
        else:
            self.banner.setText(
                "<div style='font-size:34px; font-weight:700; color:#0f4e3a;'>Sesame — Modernized</div>"
                "<div style='font-size:18px; font-weight:600; color:#1d6b55; margin-top:6px;'>"
                "Nutrient Economics for Dairy Nutritionists</div>"
            )
            self.banner.setAlignment(Qt.AlignCenter)

        self.run_button = QPushButton("Run Estimator (CSV…)")
        self.run_button.clicked.connect(self.on_run_clicked)
        self.run_button.setFixedHeight(42)
        self.run_button.setMinimumWidth(240)

        self.iterative_screening_checkbox = QCheckBox("Iterative reweighting (IRLS; Sesame 4)")
        self.iterative_screening_checkbox.setChecked(True)
        self.iterative_screening_checkbox.setToolTip(
            "Apply Sesame 4 iterative reweighted least squares: feeds are retained but downweighted based on studentized residuals."
        )
        self.iterative_screening_checkbox.stateChanged.connect(self._on_run_option_changed)

        self.save_png_checkbox = QCheckBox("Save standalone PNG files")
        self.save_png_checkbox.setChecked(False)
        self.save_png_checkbox.setToolTip(
            "When checked, SESAME will also save standalone PNG chart files in addition to embedding the figures in the Excel workbook."
        )
        self.save_png_checkbox.stateChanged.connect(self._on_run_option_changed)

        self.version_label = QLabel(f"Version {APP_VERSION}")
        self.version_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.version_label.setFont(self.subtitle_font)

        self.controls_hint = QLabel(
            "Choose a preset, then run the estimator."
        )
        self.controls_hint.setWordWrap(True)
        self.controls_hint.setAlignment(Qt.AlignCenter)
        self.controls_hint.setFont(self.small_font)

        self.hint = QLabel(
            "Use File → Run Estimator (CSV…) — each run creates a timestamped Excel workbook in /outputs"
        )
        self.hint.setAlignment(Qt.AlignLeft)
        self.hint.setWordWrap(True)
        self.hint.setFont(self.subtitle_font)

        self.selection_help = QLabel()
        self.selection_help.setWordWrap(True)
        self.selection_help.setFont(self.body_font)

        self.selection_summary = QLabel()
        self.selection_summary.setWordWrap(True)
        self.selection_summary.setFont(self.body_font)

        self.current_run_title = QLabel("Current Run Summary")
        self.current_run_title.setFont(self.header_font)
        self.current_run_title.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)

        self.current_run_summary = QLabel()
        self.current_run_summary.setWordWrap(True)
        self.current_run_summary.setTextFormat(Qt.RichText)
        self.current_run_summary.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.current_run_summary.setFont(self.body_font)

        self.current_run_footer = QLabel(
            "Updates live as you change presets, nutrient selections, or run options."
        )
        self.current_run_footer.setWordWrap(True)
        self.current_run_footer.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.current_run_footer.setFont(self.small_font)

        nutrient_group_box = QGroupBox("Nutrients to include")
        nutrient_group_box.setFont(self.header_font)
        nutrient_group_layout = QVBoxLayout()
        nutrient_group_layout.setSpacing(12)
        for group in NUTRIENT_GROUPS:
            nutrient_group_layout.addWidget(self._build_group_widget(group))
        nutrient_group_box.setLayout(nutrient_group_layout)
        self.nutrient_group_box = nutrient_group_box

    def _append_user_selected_preset(self):
        user_selected = {
            "label": "User Selected",
            "summary_label": "User Selected",
            "columns": [],
            "locked": False,
            "locked_message": "",
            "is_user_selected": True,
        }

        already_present = any(
            str(p.get("label", "")).strip().lower() == "user selected"
            for p in self._presets
        )
        if not already_present:
            self._presets.append(user_selected)

    def _build_layout(self):
        preset_row = QHBoxLayout()
        preset_row.setContentsMargins(0, 0, 0, 0)
        preset_row.setSpacing(10)
        preset_row.addWidget(self.preset_label)
        preset_row.addWidget(self.preset_box, stretch=1)

        header_layout = QHBoxLayout()
        header_layout.setContentsMargins(0, 0, 0, 0)
        header_layout.setSpacing(18)

        logo_frame = QFrame()
        logo_frame.setObjectName("headerLogoFrame")
        logo_frame.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        logo_layout = QVBoxLayout(logo_frame)
        logo_layout.setContentsMargins(14, 14, 14, 14)
        logo_layout.setSpacing(8)
        logo_layout.addStretch(1)
        logo_layout.addWidget(self.banner, alignment=Qt.AlignCenter)
        logo_layout.addStretch(1)

        control_frame = QFrame()
        control_frame.setObjectName("headerControlFrame")
        control_frame.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        control_layout = QVBoxLayout(control_frame)
        control_layout.setContentsMargins(16, 16, 16, 16)
        control_layout.setSpacing(10)
        control_layout.addWidget(self.version_label)
        control_layout.addWidget(self.run_button, alignment=Qt.AlignCenter)
        control_layout.addWidget(self.iterative_screening_checkbox, alignment=Qt.AlignCenter)
        control_layout.addWidget(self.save_png_checkbox, alignment=Qt.AlignCenter)
        control_layout.addWidget(self.controls_hint)
        control_layout.addStretch(1)

        summary_frame = QFrame()
        summary_frame.setObjectName("headerSummaryFrame")
        summary_frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        summary_layout = QVBoxLayout(summary_frame)
        summary_layout.setContentsMargins(18, 16, 18, 16)
        summary_layout.setSpacing(8)
        summary_layout.addWidget(self.current_run_title)
        summary_layout.addWidget(self.current_run_summary, stretch=1)
        summary_layout.addWidget(self.current_run_footer)

        header_layout.addWidget(logo_frame, stretch=3)
        header_layout.addWidget(control_frame, stretch=3)
        header_layout.addWidget(summary_frame, stretch=4)

        content_layout = QVBoxLayout()
        content_layout.setContentsMargins(18, 16, 18, 18)
        content_layout.setSpacing(14)
        content_layout.addLayout(preset_row)
        content_layout.addLayout(header_layout)
        content_layout.addWidget(self.hint)
        content_layout.addWidget(self.selection_help)
        content_layout.addWidget(self.selection_summary)
        content_layout.addWidget(self.nutrient_group_box)
        content_layout.addStretch(1)

        content = QWidget()
        content.setLayout(content_layout)

        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setFrameShape(QScrollArea.NoFrame)
        self.scroll_area.setWidget(content)
        self.setCentralWidget(self.scroll_area)

    def _build_menus(self):
        file_menu = self.menuBar().addMenu("&File")
        act_run = QAction("Run Estimator (CSV…)", self)
        act_run.triggered.connect(self.on_run_clicked)
        file_menu.addAction(act_run)

        file_menu.addSeparator()
        act_exit = QAction("Exit", self)
        act_exit.triggered.connect(self.close)
        file_menu.addAction(act_exit)

        about_menu = self.menuBar().addMenu("&About")
        act_about = QAction("About Sesame…", self)
        act_about.triggered.connect(self.on_about)
        about_menu.addAction(act_about)

    def _apply_global_styles(self):
        self.setStyleSheet(
            """
            QMainWindow, QWidget {
                font-size: 11pt;
            }
            QLabel {
                color: #18372d;
            }
            QLabel#muted {
                color: #4d5b57;
            }
            QGroupBox {
                font-weight: 700;
                border: 1px solid #aebbb6;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                background: #f7faf8;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 6px 0 6px;
            }
            QCheckBox {
                spacing: 8px;
                padding: 2px 0 2px 0;
                font-size: 11pt;
            }
            QComboBox {
                padding: 5px 8px;
                font-size: 11pt;
            }
            QPushButton {
                padding: 8px 14px;
                font-size: 11pt;
                font-weight: 700;
            }
            QFrame#headerLogoFrame, QFrame#headerControlFrame, QFrame#headerSummaryFrame {
                background: #f7faf8;
                border: 1px solid #c8d5cf;
                border-radius: 10px;
            }
            """
        )

        self.hint.setStyleSheet("color:#2a6f58; font-size:13pt; font-weight:700;")
        self.selection_help.setStyleSheet("color:#4d5b57; font-size:11.5pt; font-weight:600;")
        self.selection_summary.setStyleSheet("color:#143d31; font-size:12pt; font-weight:700;")
        self.version_label.setStyleSheet("color:#2a6f58;")
        self.controls_hint.setStyleSheet("color:#4d5b57;")
        self.current_run_title.setStyleSheet("color:#143d31; font-size:12pt; font-weight:700;")
        self.current_run_summary.setStyleSheet("color:#18372d; font-size:10.8pt; line-height:1.2;")
        self.current_run_footer.setStyleSheet("color:#4d5b57; font-size:10pt;")

    def _build_group_widget(self, group: Dict[str, object]) -> QGroupBox:
        box = QGroupBox(str(group["label"]))
        box.setFont(self.subtitle_font)

        grid = QGridLayout()
        grid.setHorizontalSpacing(18)
        grid.setVerticalSpacing(8)

        options = group["options"]
        for idx, option in enumerate(options):
            checkbox = QCheckBox(option["label"])
            checkbox.setFont(self.body_font)
            checkbox.stateChanged.connect(self._on_manual_selection_changed)
            self._nutrient_checkboxes[option["column"]] = checkbox
            row = idx // 2
            col = idx % 2
            grid.addWidget(checkbox, row, col)

        box.setLayout(grid)
        return box

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self._banner_pixmap:
            viewport_width = self.scroll_area.viewport().width() if hasattr(self, "scroll_area") else self.width()
            width = min(max(250, int(viewport_width * 0.22)), 350)
            height = 190
            pm = self._banner_pixmap.scaled(QSize(width, height), Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.banner.setPixmap(pm)

    def _current_preset(self) -> Dict[str, object]:
        return self._presets[self.preset_box.currentIndex()]

    def _preset_is_locked(self) -> bool:
        return bool(self._current_preset().get("locked", False))

    def _preset_is_user_selected(self) -> bool:
        return bool(self._current_preset().get("is_user_selected", False))

    def _locked_preset_message(self) -> str:
        return str(
            self._current_preset().get(
                "locked_message",
                "This preset is locked. Additional nutrient selection is disabled.",
            )
        )

    def _on_preset_changed(self, index: int):
        self._apply_preset_selection(index)
        self._refresh_checkbox_enabled_state()
        self._refresh_selection_help()
        self._refresh_selection_summary()
        self._refresh_current_run_panel()

    def _on_manual_selection_changed(self, _state: int):
        if self._manual_selection_syncing:
            return
        if self._preset_is_locked():
            self._apply_preset_selection(self.preset_box.currentIndex())
            return
        self._refresh_selection_summary()
        self._refresh_current_run_panel()

    def _on_run_option_changed(self, _state: int):
        self._refresh_current_run_panel()

    def _apply_preset_selection(self, index: int):
        preset_columns = set(self._presets[index]["columns"])
        self._manual_selection_syncing = True
        try:
            for column, checkbox in self._nutrient_checkboxes.items():
                checkbox.setChecked(column in preset_columns)
        finally:
            self._manual_selection_syncing = False

    def _refresh_checkbox_enabled_state(self):
        locked = self._preset_is_locked()
        selected_columns = set(self._current_preset()["columns"])
        for column, checkbox in self._nutrient_checkboxes.items():
            if locked:
                checkbox.setEnabled(column in selected_columns)
            else:
                checkbox.setEnabled(True)

    def _refresh_selection_help(self):
        if self._preset_is_locked():
            self.selection_help.setText(self._locked_preset_message())
        elif self._preset_is_user_selected():
            self.selection_help.setText(
                "User Selected is a manual mode. Choose the nutrients you want to include below, then run the estimator."
            )
        else:
            self.selection_help.setText(
                "Selecting a preset checks its nutrients below. You can then add or remove nutrients manually before running."
            )

    def _selected_nutrient_columns(self) -> List[str]:
        columns: List[str] = []
        for group in NUTRIENT_GROUPS:
            for option in group["options"]:
                column = option["column"]
                checkbox = self._nutrient_checkboxes[column]
                if checkbox.isChecked():
                    columns.append(column)
        return columns

    def _selected_nutrient_labels(self, columns: Optional[List[str]] = None) -> List[str]:
        cols = columns if columns is not None else self._selected_nutrient_columns()
        labels: List[str] = []
        for column in cols:
            option = NUTRIENT_OPTIONS_BY_COLUMN.get(column)
            labels.append(option["label"] if option else column)
        return labels

    def _build_effective_preset_label(self, columns: List[str]) -> str:
        preset = self._current_preset()
        preset_columns = list(preset["columns"])

        if self._preset_is_user_selected():
            return "User Selected"

        if columns == preset_columns:
            return str(preset["summary_label"])
        return f"{preset['summary_label']} + manual nutrient edits"

    def _current_output_description(self) -> str:
        if self.save_png_checkbox.isChecked():
            return "Excel workbook + embedded figures + standalone PNGs"
        return "Excel workbook + embedded figures"

    def _format_summary_list(self, values: List[str], max_items: int = 4) -> str:
        if not values:
            return "None"
        if len(values) <= max_items:
            return ", ".join(values)
        return ", ".join(values[:max_items]) + f", … (+{len(values) - max_items} more)"

    def _refresh_current_run_panel(self):
        columns = self._selected_nutrient_columns()
        labels = self._selected_nutrient_labels(columns)
        preset_label = self._build_effective_preset_label(columns)
        iterative_text = "ON" if self.iterative_screening_checkbox.isChecked() else "OFF"
        png_text = "ON" if self.save_png_checkbox.isChecked() else "OFF"

        summary_html = (
            f"<b>Version:</b> v{APP_VERSION}<br>"
            f"<b>Preset:</b> {preset_label}<br>"
            f"<b>Nutrients:</b> {len(columns)} selected<br>"
            f"<b>List:</b> {self._format_summary_list(labels)}<br>"
            f"<b>Reweighting:</b> {iterative_text}<br>"
            f"<b>PNG export:</b> {png_text}<br>"
            f"<b>Output:</b> {self._current_output_description()}<br>"
            f"<b>Folder:</b> /outputs"
        )
        self.current_run_summary.setText(summary_html)

    def _refresh_selection_summary(self):
        columns = self._selected_nutrient_columns()
        if not columns:
            self.selection_summary.setText("Currently selected nutrients: none")
            return
        self.selection_summary.setText(
            "Currently selected nutrients: " + ", ".join(self._selected_nutrient_labels(columns))
        )

    def _build_collinearity_warning(self, columns: List[str]) -> str:
        selected = set(columns)
        warnings: List[str] = []
        for required_columns, message in PROBLEMATIC_SELECTION_RULES:
            if required_columns.issubset(selected):
                warnings.append(f"• {message}")

        if len([col for col in columns if col.startswith("d") and col.endswith("_RUP")]) >= 4 and "dRUP_prot" in selected:
            warnings.append(
                "• Digestible RUP protein plus several digestible RUP amino acids may over-constrain the model in smaller feed libraries."
            )

        if not warnings:
            return ""

        return (
            "Some selected nutrients may create unstable fits or collinearity:\n\n"
            + "\n".join(warnings)
            + "\n\nThe run will continue, but review VIF values in the diagnostic workbook tab."
        )

    def on_about(self):
        QMessageBox.information(
            self,
            "About Sesame — Modernized",
            f"Sesame — Modernized\n"
            f"Version {APP_VERSION}\n"
            "Nutrient Economics for Dairy Nutritionists\n\n"
            "This software is a modern Python implementation inspired by the SESAME "
            "method described in:\n\n"
            "St-Pierre, N.R., and D. Glamocic. 2000. Estimating unit costs of nutrients "
            "from market prices of feedstuffs. Journal of Dairy Science 83:1402–1411.\n"
            "https://doi.org/10.3168/jds.S0022-0302(00)75009-0\n\n"
            "Original SESAME software is available from:\n"
            "https://dairy.osu.edu/node/23\n\n"
            "This modernized implementation extends the original framework to include "
            "iterative reweighted least squares (IRLS) to improve robustness to influential "
            "observations. This approach retains all feeds while reducing the influence of "
            "outliers through data-driven weighting.\n\n"
            "Development of the iterative reweighting approach was informed by discussions "
            "with Normand R. St-Pierre (personal communication).\n\n"
            "The default feed library distributed with this implementation originated from "
            "NASEM and may be cited as:\n"
            "NASEM. 2021. Nutrient Requirements of Dairy Cattle. Eighth revised edition. "
            "National Academies Press, Washington, DC.\n\n"
            "Modernized Python implementation developed at the University of "
            "Nebraska–Lincoln.\n\n"
            "Use File → Run Estimator (CSV…) to run an analysis.\n"
            "Outputs are written to a timestamped workbook in the /outputs folder.",
        )

    def on_run_clicked(self):
        csv_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select feed library CSV",
            self._last_dir,
            "CSV Files (*.csv);;All Files (*)",
        )
        if not csv_path:
            return
        self._last_dir = str(Path(csv_path).parent)

        cols = self._selected_nutrient_columns()
        if not cols:
            QMessageBox.warning(self, "No Nutrients Selected", "Select at least one nutrient before running the estimator.")
            return

        warning_text = self._build_collinearity_warning(cols)
        if warning_text:
            QMessageBox.warning(self, "Selection Warning", warning_text)

        starting_preset_label = str(self._current_preset()["summary_label"])
        effective_preset_label = self._build_effective_preset_label(cols)
        iterative_on = self.iterative_screening_checkbox.isChecked()
        save_standalone_pngs = self.save_png_checkbox.isChecked()

        try:
            precheck = self._estimator.summarize_input_rows(csv_path, cols)
        except Exception as e:
            tb = traceback.format_exc(limit=8)
            QMessageBox.critical(self, "Input Check Error", f"{e}\n\n{tb}")
            return

        precheck_msg = (
            f"SESAME version: {APP_VERSION}\n"
            f"Starting preset: {starting_preset_label}\n"
            f"Preset used in run: {effective_preset_label}\n"
            f"Running nutrients: {', '.join(cols)}\n"
            f"Iterative reweighting: {'ON' if iterative_on else 'OFF'}\n"
            f"{precheck['usable']} feeds usable for this nutrient set.\n"
            f"{precheck['skipped_missing_required_inputs']} feeds will be skipped due to missing required inputs."
        )
        QMessageBox.information(self, "Pre-Run Input Check", precheck_msg)

        try:
            screening = ScreeningConfig(
                enable_iterative_screening=iterative_on,
                exclude_extreme_studentized=False,
            )
            result = self._estimator.run_on_csv(csv_path, cols, screening=screening)
        except Exception as e:
            tb = traceback.format_exc(limit=8)
            QMessageBox.critical(self, "Estimator Error", f"{e}\n\n{tb}")
            return

        try:
            run_stamp = datetime.now()
            run_stamp_display = run_stamp.strftime("%Y-%m-%d %H:%M:%S")
            run_stamp_file = run_stamp.strftime("%Y-%m-%d_%H%M%S")

            run_info = self._build_run_info(
                input_csv=csv_path,
                result=result,
                selected_columns=cols,
                starting_preset_label=starting_preset_label,
                effective_preset_label=effective_preset_label,
                iterative_on=iterative_on,
                save_standalone_pngs=save_standalone_pngs,
                run_stamp_display=run_stamp_display,
                run_stamp_file=run_stamp_file,
            )

            chart_paths = self._write_chart_files(
                input_csv=csv_path,
                fit=result.final_fit,
                run_info=run_info,
                save_standalone_pngs=save_standalone_pngs,
            )

            try:
                workbook_path = self._write_outputs(
                    input_csv=csv_path,
                    result=result,
                    run_info=run_info,
                    chart_paths=chart_paths,
                )
            finally:
                self._cleanup_temp_chart_files(chart_paths)

            output_dir = Path(workbook_path).parent
            msg = self._build_run_summary(
                run_info=run_info,
                output_dir=output_dir,
                workbook_path=Path(workbook_path),
                chart_paths=chart_paths,
            )
            QMessageBox.information(self, "Run Complete", msg)

        except Exception as e:
            tb = traceback.format_exc(limit=12)
            QMessageBox.critical(self, "Write Error", f"Failed to write outputs:\n{e}\n\n{tb}")

    def _fit_rows_for_csv(self, fit: FitResult) -> List[Dict[str, object]]:
        out: List[Dict[str, object]] = []
        for row in fit.rows:
            d = _row_as_dict(row)
            actual = d.get("actual_per_t")
            pred = d.get("predicted_per_t")
            student_residual = d.get("student_residual")
            abs_student_residual = d.get("abs_student_residual")
            if abs_student_residual is None and student_residual is not None:
                abs_student_residual = abs(student_residual)

            out.append(
                {
                    "name": d.get("name", ""),
                    "actual_per_t": actual,
                    "predicted_per_t": pred,
                    "predicted_minus_actual": (pred - actual) if (pred is not None and actual is not None) else None,
                    "residual": d.get("residual"),
                    "leverage": d.get("leverage"),
                    "student_residual": student_residual,
                    "abs_student_residual": abs_student_residual,
                    "final_weight": d.get("final_weight"),
                    "excluded": False,
                    "ci75_lo": d.get("ci75_lo"),
                    "ci75_hi": d.get("ci75_hi"),
                }
            )
        return out

    def _build_run_info(
        self,
        input_csv: str,
        result: ScreeningResult,
        selected_columns: List[str],
        starting_preset_label: str,
        effective_preset_label: str,
        iterative_on: bool,
        save_standalone_pngs: bool,
        run_stamp_display: str,
        run_stamp_file: str,
    ) -> Dict[str, object]:
        input_path = Path(input_csv)

        selected_labels = self._selected_nutrient_labels(selected_columns)
        downweighted_count = sum(
            1 for row in result.final_fit.rows if getattr(row, "final_weight", 1.0) < 1.0
        )

        return {
            "sesame_version": APP_VERSION,
            "run_timestamp_display": run_stamp_display,
            "run_timestamp_file": run_stamp_file,
            "input_file_name": input_path.name,
            "input_file_stem": input_path.stem,
            "input_file_path": str(input_path.resolve()),
            "starting_preset_label": starting_preset_label,
            "effective_preset_label": effective_preset_label,
            "selected_columns": list(selected_columns),
            "selected_labels": selected_labels,
            "selected_columns_text": ", ".join(selected_columns),
            "selected_labels_text": ", ".join(selected_labels),
            "iterative_reweighting": iterative_on,
            "iterative_reweighting_text": "ON" if iterative_on else "OFF",
            "save_standalone_pngs": save_standalone_pngs,
            "save_standalone_pngs_text": "ON" if save_standalone_pngs else "OFF",
            "feeds_used_in_regression": len(result.final_fit.rows),
            "feeds_skipped_missing_inputs": len(result.pre_screen_removed_feeds),
            "feeds_excluded_final": len(result.excluded_feeds),
            "feeds_downweighted": downweighted_count,
            "final_iterations": result.final_fit.iteration_count,
            "max_iterations_reached": result.final_fit.max_iter_reached,
        }

    def _build_output_tables(self, result: ScreeningResult) -> List[Dict[str, object]]:
        final_fit = result.final_fit
        summary_rows = self._fit_rows_for_csv(final_fit)

        vif: Dict[str, float] = final_fit.vif or {}
        breakeven_rows: List[Dict[str, object]] = []
        for i, nutrient in enumerate(final_fit.nutrients):
            coef = final_fit.coef[i] if i < len(final_fit.coef) else ""
            se = final_fit.se_coef[i] if i < len(final_fit.se_coef) else ""
            breakeven_rows.append({"nutrient": nutrient, "coef": coef, "se": se, "vif": vif.get(nutrient, "")})

        breakeven_rows.extend(
            [
                {},
                {"nutrient": "adj_r2", "coef": final_fit.adj_r2},
                {"nutrient": "sigma2", "coef": final_fit.sigma2},
            ]
        )

        excluded_rows = [asdict(r) for r in result.excluded_feeds]

        report_rows: List[Dict[str, object]] = []
        excluded_map = {x.name: x.reason for x in result.excluded_feeds}
        for row in result.diagnostic_rows:
            name = str(row.get("name", ""))
            report_rows.append(
                {
                    "name": name,
                    "leverage": row.get("leverage"),
                    "student_residual": row.get("student_residual"),
                    "abs_student_residual": row.get("abs_student_residual"),
                    "final_weight": row.get("final_weight", row.get("weight", "")),
                    "excluded": name in excluded_map,
                    "exclusion_reason": excluded_map.get(name, ""),
                }
            )

        report_rows.extend(
            [
                {},
                {"name": "intercept_included_final", "leverage": result.final_fit.intercept_included},
                {"name": "intercept_pvalue_final", "leverage": result.final_fit.intercept_pvalue},
                {"name": "iteration_count_final", "leverage": result.final_fit.iteration_count},
                {"name": "max_iter_reached_final", "leverage": result.final_fit.max_iter_reached},
            ]
        )

        for nutrient, vif_value in (result.final_fit.vif or {}).items():
            concern = "ok"
            if vif_value > result.config.vif_unacceptable_threshold:
                concern = "unacceptable"
            elif vif_value > result.config.vif_concerning_threshold:
                concern = "concerning"
            report_rows.append({"name": f"vif:{nutrient}", "leverage": vif_value, "exclusion_reason": concern})

        pre_screen_rows = [
            {"name": row.get("feed_name", ""), "reason": row.get("reason", "")}
            for row in result.pre_screen_removed_feeds
        ]

        tables = [
            {
                "sheet_name": "summary",
                "fieldnames": [
                    "name",
                    "actual_per_t",
                    "predicted_per_t",
                    "predicted_minus_actual",
                    "residual",
                    "leverage",
                    "student_residual",
                    "abs_student_residual",
                    "final_weight",
                    "excluded",
                    "ci75_lo",
                    "ci75_hi",
                ],
                "rows": summary_rows,
            },
            {
                "sheet_name": "breakeven",
                "fieldnames": ["nutrient", "coef", "se", "vif"],
                "rows": breakeven_rows,
            },
            {
                "sheet_name": "initial_fit_summary",
                "fieldnames": [
                    "name",
                    "actual_per_t",
                    "predicted_per_t",
                    "predicted_minus_actual",
                    "residual",
                    "leverage",
                    "student_residual",
                    "abs_student_residual",
                    "final_weight",
                    "excluded",
                    "ci75_lo",
                    "ci75_hi",
                ],
                "rows": self._fit_rows_for_csv(result.initial_fit),
            },
            {
                "sheet_name": "final_fit_summary",
                "fieldnames": [
                    "name",
                    "actual_per_t",
                    "predicted_per_t",
                    "predicted_minus_actual",
                    "residual",
                    "leverage",
                    "student_residual",
                    "abs_student_residual",
                    "final_weight",
                    "excluded",
                    "ci75_lo",
                    "ci75_hi",
                ],
                "rows": summary_rows,
            },
            {
                "sheet_name": "excluded_feeds",
                "fieldnames": ["name", "reason", "leverage", "student_residual"],
                "rows": excluded_rows,
            },
            {
                "sheet_name": "diagnostic_report",
                "fieldnames": [
                    "name",
                    "leverage",
                    "student_residual",
                    "abs_student_residual",
                    "final_weight",
                    "excluded",
                    "exclusion_reason",
                ],
                "rows": report_rows,
            },
            {
                "sheet_name": "pre_screen_removed_feeds",
                "fieldnames": ["name", "reason"],
                "rows": pre_screen_rows,
            },
        ]

        if result.iteration_log:
            tables.append(
                {
                    "sheet_name": "iteration_log",
                    "fieldnames": [
                        "iteration",
                        "feed_name",
                        "weight",
                        "residual",
                        "student_residual",
                        "abs_student_residual",
                    ],
                    "rows": result.iteration_log,
                }
            )

        return tables

    def _write_outputs(
        self,
        input_csv: str,
        result: ScreeningResult,
        run_info: Dict[str, object],
        chart_paths: Dict[str, Optional[str]],
    ) -> str:
        inp = Path(input_csv)
        out_dir = self._ensure_output_dir()
        workbook_path = out_dir / f"{inp.stem}_{run_info['run_timestamp_file']}.results.xlsx"
        tables = self._build_output_tables(result)

        self._write_excel_workbook(
            workbook_path=workbook_path,
            tables=tables,
            run_info=run_info,
            chart_paths=chart_paths,
        )
        return str(workbook_path)

    def _write_excel_workbook(
        self,
        workbook_path: Path,
        tables: List[Dict[str, object]],
        run_info: Dict[str, object],
        chart_paths: Dict[str, Optional[str]],
    ):
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except Exception as e:
            raise RuntimeError(
                f"Excel export requires openpyxl and pillow. Import failed: {e}"
            )

        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

        header_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
        title_fill = PatternFill(fill_type="solid", fgColor="B6D7A8")
        header_font = Font(bold=True)
        title_font = Font(bold=True, size=12)

        ws_info = wb.create_sheet(title="Run Info")
        ws_info["A1"] = "SESAME Run Information"
        ws_info["A1"].font = title_font
        ws_info["A1"].fill = title_fill

        run_info_rows = [
            ("SESAME version", run_info["sesame_version"]),
            ("Run timestamp", run_info["run_timestamp_display"]),
            ("Input file name", run_info["input_file_name"]),
            ("Input file path", run_info["input_file_path"]),
            ("Starting preset", run_info["starting_preset_label"]),
            ("Preset used", run_info["effective_preset_label"]),
            ("Selected nutrients (columns)", run_info["selected_columns_text"]),
            ("Selected nutrients (labels)", run_info["selected_labels_text"]),
            ("Iterative reweighting", run_info["iterative_reweighting_text"]),
            ("Save standalone PNG files", run_info["save_standalone_pngs_text"]),
            ("Feeds used in regression", run_info["feeds_used_in_regression"]),
            ("Feeds skipped due to missing inputs", run_info["feeds_skipped_missing_inputs"]),
            ("Feeds excluded in final output", run_info["feeds_excluded_final"]),
            ("Feeds downweighted", run_info["feeds_downweighted"]),
            ("Final iterations", run_info["final_iterations"]),
            ("Max iterations reached", run_info["max_iterations_reached"]),
        ]

        for r_idx, (label, value) in enumerate(run_info_rows, start=3):
            ws_info.cell(row=r_idx, column=1, value=label)
            ws_info.cell(row=r_idx, column=2, value=value)
            ws_info.cell(row=r_idx, column=1).font = header_font
            ws_info.cell(row=r_idx, column=1).fill = header_fill
            ws_info.cell(row=r_idx, column=1).alignment = Alignment(vertical="top")
            ws_info.cell(row=r_idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")

        ws_info.column_dimensions["A"].width = 34
        ws_info.column_dimensions["B"].width = 110
        ws_info.freeze_panes = "A3"

        ws_fig = wb.create_sheet(title="Figures")
        ws_fig["A1"] = "SESAME Figures"
        ws_fig["A1"].font = title_font
        ws_fig["A1"].fill = title_fill
        ws_fig["A3"] = f"SESAME version: {run_info['sesame_version']}"
        ws_fig["A4"] = f"Preset used: {run_info['effective_preset_label']}"
        ws_fig["A5"] = f"Nutrients: {run_info['selected_labels_text']}"
        ws_fig["A6"] = f"Iterative reweighting: {run_info['iterative_reweighting_text']}"
        ws_fig["A7"] = f"Input file: {run_info['input_file_name']}"
        ws_fig["A8"] = f"Run timestamp: {run_info['run_timestamp_display']}"

        for col in "ABCDEFGHIJKL":
            ws_fig.column_dimensions[col].width = 22

        chart_embed_path = chart_paths.get("chart_embed")
        opp_embed_path = chart_paths.get("opportunity_embed")

        if chart_embed_path and Path(chart_embed_path).exists():
            img1 = XLImage(chart_embed_path)
            self._scale_excel_image(img1, max_width=1450, max_height=720)
            ws_fig.add_image(img1, "A10")

        if opp_embed_path and Path(opp_embed_path).exists():
            img2 = XLImage(opp_embed_path)
            self._scale_excel_image(img2, max_width=1450, max_height=880)
            ws_fig.add_image(img2, "A45")

        used_sheet_names = {"Run Info", "Figures"}
        for table in tables:
            sheet_name = self._unique_sheet_name(str(table["sheet_name"]), used_sheet_names)
            ws = wb.create_sheet(title=sheet_name)
            fieldnames = list(table["fieldnames"])
            rows = list(table["rows"])

            for col_idx, field in enumerate(fieldnames, start=1):
                cell = ws.cell(row=1, column=col_idx, value=field)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

            for row_idx, row in enumerate(rows, start=2):
                for col_idx, field in enumerate(fieldnames, start=1):
                    value = row.get(field, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            ws.freeze_panes = "A2"

            for col_idx, field in enumerate(fieldnames, start=1):
                longest = len(str(field))
                for row in rows:
                    value = row.get(field, "")
                    if value is None:
                        value = ""
                    longest = max(longest, len(str(value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(longest + 2, 12), 40)

        wb.save(workbook_path)

    def _unique_sheet_name(self, name: str, used_names: set[str]) -> str:
        sanitized = re.sub(r"[\\/*?:\[\]]", "_", name).strip() or "sheet"
        sanitized = sanitized[:31]
        candidate = sanitized
        counter = 2
        while candidate in used_names:
            suffix = f"_{counter}"
            candidate = f"{sanitized[:31 - len(suffix)]}{suffix}"
            counter += 1
        used_names.add(candidate)
        return candidate

    def _scale_excel_image(self, img, max_width: int, max_height: int):
        original_width = img.width
        original_height = img.height

        if not original_width or not original_height:
            return

        scale = min(max_width / original_width, max_height / original_height, 1.0)
        img.width = int(original_width * scale)
        img.height = int(original_height * scale)

    def _ensure_output_dir(self) -> Path:
        out_dir = self._project_root / "outputs"
        out_dir.mkdir(parents=True, exist_ok=True)
        return out_dir

    def _build_chart_metadata_lines(self, run_info: Dict[str, object]) -> List[str]:
        labels = list(run_info["selected_labels"])
        if len(labels) <= 6:
            nutrient_text = ", ".join(labels)
        else:
            nutrient_text = ", ".join(labels[:6]) + f", … (+{len(labels) - 6} more)"

        return [
            f"SESAME v{run_info['sesame_version']} | Preset: {run_info['effective_preset_label']}",
            f"Nutrients: {nutrient_text}",
            f"Iterative reweighting: {run_info['iterative_reweighting_text']} | Input: {run_info['input_file_name']} | Run: {run_info['run_timestamp_display']}",
        ]

    def _write_chart_files(
        self,
        input_csv: str,
        fit: FitResult,
        run_info: Dict[str, object],
        save_standalone_pngs: bool,
    ) -> Dict[str, Optional[str]]:
        chart_info = self._write_bar_chart(
            input_csv=input_csv,
            fit=fit,
            run_info=run_info,
            save_standalone_png=save_standalone_pngs,
        )
        opportunity_info = self._write_opportunity_plot(
            input_csv=input_csv,
            fit=fit,
            run_info=run_info,
            save_standalone_png=save_standalone_pngs,
        )
        return {
            "chart_embed": chart_info["embed"],
            "chart_standalone": chart_info["standalone"],
            "opportunity_embed": opportunity_info["embed"],
            "opportunity_standalone": opportunity_info["standalone"],
        }

    def _cleanup_temp_chart_files(self, chart_paths: Dict[str, Optional[str]]):
        for key in ("chart_embed", "opportunity_embed"):
            path = chart_paths.get(key)
            if path:
                try:
                    p = Path(path)
                    if p.exists():
                        p.unlink()
                except Exception:
                    pass

    def _write_bar_chart(
        self,
        input_csv: str,
        fit: FitResult,
        run_info: Dict[str, object],
        save_standalone_png: bool,
    ) -> Dict[str, Optional[str]]:
        try:
            import matplotlib.pyplot as plt
            import numpy as np
        except Exception as e:
            raise RuntimeError(f"Matplotlib import failed: {e}")

        inp = Path(input_csv)
        out_dir = self._ensure_output_dir()

        names = [r.name for r in fit.rows]
        actual = np.array([r.actual_per_t for r in fit.rows], dtype=float)
        predicted = np.array([r.predicted_per_t for r in fit.rows], dtype=float)
        value = predicted - actual

        order = np.argsort(value)
        names = [names[i] for i in order]
        actual = actual[order]
        predicted = predicted[order]
        value = value[order]

        x = np.arange(len(names))
        width = 0.28

        fig_w = max(14, min(34, 0.60 * len(names)))
        fig_h = 10.5
        fig, ax = plt.subplots(figsize=(fig_w, fig_h))

        ax.bar(x - width, actual, width, label="Actual ($/t)")
        ax.bar(x, predicted, width, label="Predicted ($/t)")
        ax.bar(x + width, value, width, label="Predicted − Actual ($/t)")

        ax.axhline(0, linewidth=1)
        ax.set_xticks(x)
        ax.set_xticklabels(names, rotation=52, ha="right", fontsize=10)
        ax.set_ylabel("$/ton", fontsize=11)
        ax.set_title("Sesame Feed Value Chart", fontsize=22, fontweight="bold", pad=24)
        ax.legend(fontsize=11, loc="best")
        ax.tick_params(axis="y", labelsize=10)

        meta_lines = self._build_chart_metadata_lines(run_info)
        fig.text(0.01, 0.975, meta_lines[0], fontsize=10, ha="left", va="top")
        fig.text(0.01, 0.953, meta_lines[1], fontsize=9, ha="left", va="top")
        fig.text(0.01, 0.931, meta_lines[2], fontsize=9, ha="left", va="top")

        fig.tight_layout(rect=[0.02, 0.18, 0.995, 0.91])

        embed_path = out_dir / f"{inp.stem}_{run_info['run_timestamp_file']}.chart_embed.png"
        fig.savefig(embed_path, dpi=220, bbox_inches="tight")

        standalone_path: Optional[Path] = None
        if save_standalone_png:
            standalone_path = out_dir / f"{inp.stem}_{run_info['run_timestamp_file']}.chart.png"
            fig.savefig(standalone_path, dpi=220, bbox_inches="tight")

        plt.close(fig)

        return {
            "embed": str(embed_path),
            "standalone": str(standalone_path) if standalone_path else None,
        }

    def _write_opportunity_plot(
        self,
        input_csv: str,
        fit: FitResult,
        run_info: Dict[str, object],
        save_standalone_png: bool,
    ) -> Dict[str, Optional[str]]:
        try:
            import matplotlib.pyplot as plt
            import numpy as np
        except Exception as e:
            raise RuntimeError(f"Matplotlib import failed: {e}")

        inp = Path(input_csv)
        out_dir = self._ensure_output_dir()

        names = [r.name for r in fit.rows]
        predicted = np.array([r.predicted_per_t for r in fit.rows], dtype=float)
        actual = np.array([r.actual_per_t for r in fit.rows], dtype=float)
        value = predicted - actual

        fig, ax = plt.subplots(figsize=(13, 10))

        ax.scatter(predicted, value, s=42)

        for i, name in enumerate(names):
            ax.annotate(
                name,
                (predicted[i], value[i]),
                textcoords="offset points",
                xytext=(5, 4),
                ha="left",
                fontsize=8,
            )

        ax.axhline(0, linewidth=1)
        ax.set_xlabel("Predicted price ($/t)", fontsize=11)
        ax.set_ylabel("Predicted − Actual ($/t)  (positive = undervalued)", fontsize=11)
        ax.set_title("Sesame Opportunity Plot", fontsize=20, fontweight="bold", pad=24)
        ax.tick_params(axis="both", labelsize=10)

        meta_lines = self._build_chart_metadata_lines(run_info)
        fig.text(0.01, 0.975, meta_lines[0], fontsize=10, ha="left", va="top")
        fig.text(0.01, 0.953, meta_lines[1], fontsize=9, ha="left", va="top")
        fig.text(0.01, 0.931, meta_lines[2], fontsize=9, ha="left", va="top")

        fig.tight_layout(rect=[0.03, 0.04, 0.995, 0.91])

        embed_path = out_dir / f"{inp.stem}_{run_info['run_timestamp_file']}.opportunity_embed.png"
        fig.savefig(embed_path, dpi=220, bbox_inches="tight")

        standalone_path: Optional[Path] = None
        if save_standalone_png:
            standalone_path = out_dir / f"{inp.stem}_{run_info['run_timestamp_file']}.opportunity.png"
            fig.savefig(standalone_path, dpi=220, bbox_inches="tight")

        plt.close(fig)

        return {
            "embed": str(embed_path),
            "standalone": str(standalone_path) if standalone_path else None,
        }

    def _build_run_summary(
        self,
        run_info: Dict[str, object],
        output_dir: Path,
        workbook_path: Path,
        chart_paths: Dict[str, Optional[str]],
    ) -> str:
        msg = (
            f"SESAME version: {run_info['sesame_version']}\n"
            f"Preset used: {run_info['effective_preset_label']}\n"
            f"Nutrients run: {run_info['selected_columns_text']}\n"
            f"Iterative reweighting: {run_info['iterative_reweighting_text']}\n"
            f"• Feeds used in regression: {run_info['feeds_used_in_regression']}\n"
            f"• Feeds skipped due to missing inputs: {run_info['feeds_skipped_missing_inputs']}\n"
            f"• Feeds downweighted by iterative reweighting: {run_info['feeds_downweighted']}\n"
            f"• Final iterations: {run_info['final_iterations']}\n"
            f"• Max iterations reached: {run_info['max_iterations_reached']}\n"
            f"• Workbook created: {workbook_path.name}\n"
        )

        if run_info["save_standalone_pngs"]:
            chart_name = Path(chart_paths["chart_standalone"]).name if chart_paths.get("chart_standalone") else ""
            opp_name = Path(chart_paths["opportunity_standalone"]).name if chart_paths.get("opportunity_standalone") else ""
            msg += (
                f"• Standalone chart PNG: {chart_name}\n"
                f"• Standalone opportunity PNG: {opp_name}\n"
            )
        else:
            msg += "• Figures embedded in workbook only (standalone PNGs not saved)\n"

        msg += f"\nOutput files saved in: {output_dir}"
        return msg